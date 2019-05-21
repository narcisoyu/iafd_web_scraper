import requests
from lxml import etree
from openpyxl import Workbook

all_data1 = []
all_data2 = []
HEADERS = {
            'Cookie': 'ASPSESSIONIDCCTCSBCQ=BMBKGPFAOPDNHJNJIDBGIJPE',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.103 Safari/537.36'
            }
all_index = ['http://iafd.com/astrology.rme/sign=Aries',
            'http://iafd.com/astrology.rme/sign=Taurus',
            'http://iafd.com/astrology.rme/sign=Gemini',
            'http://iafd.com/astrology.rme/sign=Cancer',
            'http://iafd.com/astrology.rme/sign=Leo',
            'http://iafd.com/astrology.rme/sign=Virgo',
            'http://iafd.com/astrology.rme/sign=Libra',
            'http://iafd.com/astrology.rme/sign=Scorpio',
            'http://iafd.com/astrology.rme/sign=Sagittarius',
            'http://iafd.com/astrology.rme/sign=Capricorn',
            'http://iafd.com/astrology.rme/sign=Aquarius',
            'http://iafd.com/astrology.rme/sign=Pisces']

for index in all_index:
    try:
        response = requests.get(url = index,headers = HEADERS)
        root = etree.HTML(response.content)
        person_url = root.xpath('//*[@id="astro"]/div/a/@href')
    except:
        continue
    for url in person_url[:3]:
        url = 'http://iafd.com'+ url
        print(url)
        try:
            person_response = requests.get(url = url,headers = HEADERS, timeout = 20)
        except:
            continue
        person_root = etree.HTML(person_response.content)

        try:
            name = person_root.xpath('/html/body/div/div[1]/div/h1/text()')[0]
        except:
            name = ''
        try:
            ETHNICITY = person_root.xpath('//*[@id="home"]/div/div[1]/p[2]/text()')[0]
        except:
            ETHNICITY = ''
        try:
            HEIGHT = person_root.xpath('//*[@id="home"]/div/div[2]/p[2]/text()')[0]
        except:
            HEIGHT = ''
        try:
            TATTOOS = person_root.xpath('//*[@id="home"]/div/div[3]/p[2]/text()')[0]
        except:
            TATTOOS = ''
        try:
            NATIONALITY = person_root.xpath('//*[@id="home"]/div/div[1]/p[4]/text()')[0]
        except:
            NATIONALITY = ''
        try:
            WEIGHT = person_root.xpath('//*[@id="home"]/div/div[2]/p[4]/text()')[0]
        except:
            WEIGHT = ''
        try:
            PIERCINGS = person_root.xpath('//*[@id="home"]/div/div[3]/p[4]/text()')[0]
        except:
            PIERCINGS = ''
        try:
            HAIR_COLOR = person_root.xpath('//*[@id="home"]/div/div[1]/p[6]/text()')[0]
        except:
            HAIR_COLOR = ''
        try:
            MEASUREMENTS = person_root.xpath('//*[@id="home"]/div/div[2]/p[6]/text()')[0]
        except:
            MEASUREMENTS = ''

        data = []
        data.append(name)
        data.append(ETHNICITY)
        data.append(HEIGHT)
        data.append(TATTOOS)
        data.append(PIERCINGS)
        data.append(HAIR_COLOR)
        data.append(MEASUREMENTS)
        all_data1.append(data)
        for index,value in enumerate(person_root.xpath('//*[@id="personal"]/tbody/tr/td[1]/a')):
            try:
                Movie_title = person_root.xpath('//*[@id="personal"]/tbody/tr['+str(index+1)+']/td[1]/a/text()')[0]
            except:
                Movie_title=''
            try:
                Year = person_root.xpath('//*[@id="personal"]/tbody/tr['+str(index+1)+']/td[2]/a/text()')[0]
            except:
                Year = ''
            try:
                Distributor = person_root.xpath('//*[@id="personal"]/tbody/tr['+str(index+1)+']/td[3]/a/text()')[0]
            except:
                Distributor=''
            try:
                Notes = person_root.xpath('//*[@id="personal"]/tbody/tr['+str(index+1)+']/td[4]/a/text()')[0]
            except:
                Notes = ''
            try:
                Rev = person_root.xpath('//*[@id="personal"]/tbody/tr['+str(index+1)+']/td[5]/a/text()')[0]
            except:
                Rev = ''
            try:
                Formats = person_root.xpath('//*[@id="personal"]/tbody/tr['+str(index+1)+']/td[6]/a/text()')[0]
            except:
                Formats = ''

            data = []
            data.append(name)
            data.append(Movie_title)
            data.append(Year)
            data.append(Distributor)
            data.append(Notes)
            data.append(Rev)
            data.append(Formats)

            all_data2.append(data)


wb = Workbook()  # 新建一个excel
sheet = wb.create_sheet()
row = 1  # 控制行
for data in all_data1:
    col = 1  # 控制列
    for s in data:  # 再循环里面list的值，每一列
        s = str(s)
        s.replace('\r', '')
        s.replace('\n', '')

        sheet.cell(row=row, column=col, value=s)
        col += 1
    row += 1

wb.save('actor_data.xlsx')


wb2 = Workbook()  # 新建一个excel
sheet2 = wb2.create_sheet()
row = 1  # 控制行
for data in all_data2:
    col = 1  # 控制列
    for s in data:  # 再循环里面list的值，每一列
        s = str(s)
        s.replace('\r', '')
        s.replace('\n', '')

        sheet2.cell(row=row, column=col, value=s)
        col += 1
    row += 1

wb2.save('movie_Data.xlsx')